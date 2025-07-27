from datetime import datetime

from celery import shared_task

from colleges.models import College
from colleges.reports import get_mobility_report, get_performance_report
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMessage

@shared_task
def generate_institutional_report(college_id, start_date_str, end_date_str, email_to):
    college = College.objects.get(pk=college_id)
    start_date = datetime.fromisoformat(start_date_str)
    end_date = datetime.fromisoformat(end_date_str)

    mobility = get_mobility_report(college, start_date, end_date)
    performance = get_performance_report(college, start_date, end_date)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(fill_type="solid", start_color="722352", end_color="722352")

    ws1 = wb.active
    ws1.title = "Movilidad"

    # Title for Mobility
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = ws1.cell(row=1, column=1)
    title_cell.value = "Resumen de Movilidad"
    title_cell.font = Font(bold=True, size=14)
    headers1 = ['Tipo', 'ID', 'Nombre de usuario', 'Total Viajes']
    ws1.append(headers1)

    for cell in ws1[ws1.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    # Students
    for item in mobility['student_trip_counts']:
        ws1.append(['Estudiante', item['passenger'], item['passenger__username'], item['count']])

    # Drivers
    for item in mobility['driver_trip_counts']:
        ws1.append(['Conductor', item['driver'], item['driver__username'], item['count']])

    # Vehicles
    for item in mobility['vehicle_trip_counts']:
        ws1.append(['Vehículo', item['vehicle'], 'No aplica', item['count']])

    # Top Hours
    ws1.append([])
    ws1.merge_cells(start_row=ws1.max_row + 1, start_column=1, end_row=ws1.max_row + 1, end_column=2)
    title_cell = ws1.cell(row=ws1.max_row + 1, column=1)
    title_cell.value = "Horas con mayor demanda"
    title_cell.font = Font(bold=True, size=14)
    headers2 = ['Hora', 'Viajes']
    ws1.append(headers2)
    for cell in ws1[ws1.max_row]:
        cell.fill = header_fill
        cell.font = header_font
    for hour, cnt in mobility['top_hours']:
        ws1.append([hour, cnt])

    # Routes
    ws1.append([])
    ws1.merge_cells(start_row=ws1.max_row + 1, start_column=1, end_row=ws1.max_row + 1, end_column=2)
    title_cell = ws1.cell(row=ws1.max_row + 1, column=1)
    title_cell.value = "Rutas más utilizadas"
    title_cell.font = Font(bold=True, size=14)
    headers3 = ['ID', 'Ruta', 'Viajes']
    ws1.append(headers3)
    for cell in ws1[ws1.max_row]:
        cell.fill = header_fill
        cell.font = header_font
    for item in mobility['top_routes']:
        ws1.append([item['route'], item['route__name'], item['count']])

    # 2nd Sheet: Security
    ws2 = wb.create_sheet('Desempeño')
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = ws2.cell(row=1, column=1)
    title_cell.value = "Comentarios de Pasajeros"
    title_cell.font = Font(bold=True, size=14)
    # Feedback
    headers4 = ['Pasajero', 'Comentario']
    ws2.append(headers4)
    for cell in ws2[ws2.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    for fb in performance['passenger_feedback']:
        ws2.append([fb['passenger'], fb['comment']])

    # Ratings
    ws2.append([])
    ws2.merge_cells(start_row=ws2.max_row + 1, start_column=1, end_row=ws2.max_row + 1, end_column=3)
    title_cell = ws2.cell(row=ws2.max_row + 1, column=1)
    title_cell.value = "Calificaciones de Conductores"
    title_cell.font = Font(bold=True, size=14)
    headers5 = ['ID', 'Nombre de Usuario', 'Calificación', 'Total']
    ws2.append(headers5)
    for cell in ws2[ws2.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    for rt in performance['driver_avg_ratings']:
        ws2.append([rt['trip__driver'], rt['trip__driver__username'], rt['avg_ratting'], rt['total']])

    # 3rd sheet: Security
    ws3 = wb.create_sheet('Seguridad')
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    title_cell = ws3.cell(row=1, column=1)
    title_cell.value = "Indicadores de Seguridad"
    title_cell.font = Font(bold=True, size=14)
    headers6 = ['Total Incidentes', 'Total Desvíos', 'Total Viajes']
    ws3.append(headers6)
    for cell in ws3[ws3.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    ws3.append([
        performance['total_incidents'],
        performance['total_deviations'],
        performance['total_trips']
    ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    subject = f"Reporte Institucional {college.name}: {start_date} a {end_date}"
    body = render_to_string('emails/college_report_ready.txt', {
        'college_name': college.name,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d')
    })
    email = EmailMessage(
        subject,
        body,
        from_email=settings.EMAIL_HOST_USER,
        to=[email_to]
    )
    filename = f"college_{college.college_id}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    email.attach(filename, buf.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send(fail_silently=False)